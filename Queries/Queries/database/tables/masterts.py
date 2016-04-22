import os
import logging
import datetime
from   openpyxl  import load_workbook
#import timesheet as     TimeSheet

#-----------------------------------------------------------------------
class Master:
  data = None

  #---------------------------------------------------------------------
  # Empty class except that it reads a spreadsheet and populates
  # other static variables.
  #---------------------------------------------------------------------
  def __init__(self,root):
    masterSS = None
    candidateList = []
    for file in os.listdir(root):
      #logging.debug(file)
      if file.endswith('xlsx'):
        if file.startswith('FAE Timesheet'):
          candidateList.append(file)
    
    if (len(candidateList) == 0):
      raise Exception
    
    masterSS = Master.getLastestMaster(candidateList)

    # this is bad and has the side effect of initializing
    # Timesheet field classes, but will do for now.
    self.data = Master.read(os.path.join(root,masterSS))

  #-----------------------------------------------------------------------
  def getLastestMaster(list):

    #---------------------------------------------------------------------
    class FileName:
      def __init__(self,fname):
        self.fname = fname
        list = fname.split()
        self.fae   = list[0].strip()
        self.ts    = list[1].strip()
        ver        = list[2].strip()
        ver        = ver.lower();
        self.ver   = float(ver.lstrip('v'))
        date       = list[3].strip()
        day        = int(date[0:2])
        mon        = int(date[3:5])
        year       = int(date[6:8])
        self.date  = datetime.date(year,mon,day)
        self.ext   = list[3].strip()
        self.ext   = self.ext[9:13]
        # should error check ver, date, ext)
        #logging.debug('|' + str(self.ver) +'|' + str(self.date) + '|')

      def __lt__(self,other):
        if (self.ver < other.ver):
          return True
        else:
          return False      
      def __gt__(self,other):
        if (self.ver > other.ver):
          return True
        else:
          return False

    #---------------------------------------------------------------------
    fnameList = []

    for i,file in enumerate(list):
      filename = FileName(file)
      if (filename.date != None and filename.ver != None):
        fnameList.append(filename)

    fnameList.sort(reverse=True)
    return fnameList[0].fname

  #-----------------------------------------------------------------------
  # Read codes, locations, activities, products
  #-----------------------------------------------------------------------
  def read(fname):
    wb = load_workbook(fname)

    sheet_names = wb.get_sheet_names();
    if (sheet_names[0] != 'Timesheet'):
      raise Exception

    ws = wb.get_sheet_by_name('Timesheet')

    codes      = Master.readCol(ws, 5)
    locations  = Master.readCol(ws, 6)
    activities = Master.readCol(ws, 7)
    products   = Master.readCol(ws, 8)

    results = {}
    results['CODES'     ] = codes
    results['LOCATIONS' ] = locations
    results['ACTIVITIES'] = activities
    results['PRODUCTS'  ] = products

    return results

  #-----------------------------------------------------------------------
  def readCol(ws,col):
    list = None
    startRow = 50
    found    = False
    endCnt   = 0

    #---------------------------------------------------------------------
    # Loop looking for values and saving them
    #---------------------------------------------------------------------
    wsRow = startRow
    wsCol =  col
    while True:
      val = ws.cell(row = wsRow,column=wsCol).value
      if (val != None):
        if (found == False):
          found = True
          startRow = wsRow
          list = []
        list.append(val)
        endCnt = 0
        #logging.debug(val)
      else:
        if (found != False):
          endCnt += 1

      if (endCnt > 5 or wsRow > 500):
        break;

      wsRow += 1

    if (list == None or len(list) == 0):
      raise Exception

    return list


