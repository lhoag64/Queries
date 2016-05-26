import logging
from   xlinterface.xlworkbook            import XlWorkBook
from   xlinterface.xlworksheet           import XlWorkSheet
from   summary.matrix.matrixdata         import MatrixData 
from   summary.matrix.matrixtable        import MatrixTable
from   openpyxl.chart                    import LineChart
from   openpyxl.chart                    import BarChart
from   openpyxl.chart                    import PieChart
from   openpyxl.chart.reference          import Reference
from   summary.matrix.actvitivydata      import ActivityData
from   summary.matrix.ltsdata            import LtsData
from   summary.matrix.gkadata            import GkaData
from   summary.matrix.actbylocdata       import ActByLocData

#----------------------------------------------------------------------
class MetricSummarySheet:
  #--------------------------------------------------------------------
  def __init__(self,ws,region,type,period):
    self.ws     = ws
    self.region = region
    self.type   = type
    self.period = period

  #--------------------------------------------------------------------
  def drawSummary(self):
    self.ws.SetColWid(2,10)
    self.ws.SetColWid(3,60)
    self.ws.SetColWid(4,20)
    self.ws.SetColWid(5,20)
    self.ws.SetColWid(6,10)

    self.ws.DrawRegion(2,2,66,6,'thick','Blue 1')

    self.ws.ws.merge_cells('C3:E3')

    fmt1  = {'hAlign':'C','vAlign':'C','font':{'emph':'B','size':14}}
    fmt2  = {'hAlign':'L','vAlign':'C','border':{'A':'thin'},'fill':'Blue 2','font':{'emph':'B','size':11}}
    fmt2a = {'hAlign':'C','vAlign':'C','border':{'A':'thin'},'fill':'Blue 2','font':{'emph':'B','size':11}}
    fmt2b = {'hAlign':'C','vAlign':'C','border':{'A':'thin'},'fill':'Blue 2','numFmt':'0.0','font':{'emph':'B'}}
    fmt3  = {'hAlign':'L','vAlign':'C','border':{'A':'thin'},'fill':'Blue 3'}
    fmt3a = {'hAlign':'C','vAlign':'C','border':{'A':'thin'},'fill':'Blue 3'}
    fmt3b = {'hAlign':'C','vAlign':'C','border':{'A':'thin'},'fill':'Blue 3','numFmt':'0.0','font':{'emph':'B'}}
    fmt4  = {'hAlign':'L','vAlign':'C','border':{'A':'thin'},'fill':'Green 1'}
    fmt4b = {'hAlign':'C','vAlign':'C','border':{'A':'thin'},'fill':'Green 1','numFmt':'0.0','font':{'emph':'B'}}
    fmt5  = {'hAlign':'L','vAlign':'C','border':{'A':'thin'},'fill':'Orange 1'}
    fmt5b = {'hAlign':'C','vAlign':'C','border':{'A':'thin'},'fill':'Orange 1','numFmt':'0.0','font':{'emph':'B'}}
    fmt6  = {'hAlign':'L','vAlign':'C','border':{'A':'thin'},'fill':'Red 1'}
    fmt6b = {'hAlign':'C','vAlign':'C','border':{'A':'thin'},'fill':'Red 1','numFmt':'0.0','font':{'emph':'B'}}
    fmt7  = {'hAlign':'L','vAlign':'C','border':{'A':'thin'},'fill':'Yellow 1'}
    fmt7b = {'hAlign':'C','vAlign':'C','border':{'A':'thin'},'fill':'Yellow 1','numFmt':'0.0','font':{'emph':'B'}}

    self.ws.SetCell( 3, 3,'Monthly Summary Statement: Year To Date (YTD)',fmt1)

    self.ws.SetCell( 6, 3,'Hours',fmt2)
    self.ws.SetCell( 6, 4,'Hours',fmt2a)
    self.ws.SetCell( 6, 5,'As a % of Contracted',fmt2a)
    self.ws.SetCell( 7, 3,'Contracted number of hours',fmt3)
    self.ws.SetCell( 7, 4,'=SUM(GS243:HE243)',fmt3b)
    self.ws.SetCell( 7, 5,None,fmt3b)
    self.ws.SetCell( 8, 3,'Total Hours booked',fmt3)
    self.ws.SetCell( 8, 4,'=SUM(GS223:HE223)',fmt3b)
    self.ws.SetCell( 8, 5,None,fmt3b)
    self.ws.SetCell( 9, 3,'Additional hours worked over contracted',fmt3)
    self.ws.SetCell( 9, 4,'=SUM(GS242:HE242)',fmt3b)
    self.ws.SetCell( 9, 5,None,fmt3b)
    self.ws.SetCell(10, 3,'Number of Heads',fmt3)
    self.ws.ws.merge_cells('D10:E10')
    self.ws.SetCell(10, 4,16,fmt3b)
    self.ws.SetCell(10, 5,None,fmt3b)

    self.ws.ws.merge_cells('D12:E12')
    self.ws.ws.merge_cells('D13:E13')
    self.ws.ws.merge_cells('D14:E14')
    self.ws.ws.merge_cells('D15:E15')
    self.ws.ws.merge_cells('D16:E16')
    self.ws.SetCell(12, 3,'Activity (Summary) AVERAGE ACCROSS YEAR',fmt2)
    self.ws.SetCell(12, 4,'Percentage %',fmt2a)
    self.ws.SetCell(13, 3,'Utilisation (Customer Funded works)',fmt4)
    self.ws.SetCell(13, 4,'=HF224',fmt4b)
    self.ws.SetCell(13, 5,None,fmt4b)
    self.ws.SetCell(14, 3,'Utilisation (Pre Sales work)',fmt5)
    self.ws.SetCell(14, 4,'=HF229',fmt5b)
    self.ws.SetCell(14, 5,None,fmt5b)
    self.ws.SetCell(15, 3,'Utilisation (Downtime,Exc Leave and Sickness)',fmt6)
    self.ws.SetCell(15, 4,'=HF234',fmt6b)
    self.ws.SetCell(15, 5,None,fmt6b)
    self.ws.SetCell(16, 3,'Utilisation (Leave and Sickness)',fmt7)
    self.ws.SetCell(16, 4,'=HF239',fmt7b)
    self.ws.SetCell(16, 5,None,fmt7b)

    self.ws.SetCell(19, 3,'Activity (Detailed)',fmt2)
    self.ws.SetCell(19, 4,'Hours',fmt2a)
    self.ws.SetCell(19, 5,'As a % of Total',fmt2a)
    self.ws.SetCell(20, 3,'Support agreement (Software)',fmt4)
    self.ws.SetCell(20, 4,'=SUM(GS201:HE201)',fmt4b)
    self.ws.SetCell(20, 5,'=D20/SUM(D20:D27)*100',fmt4b)
    self.ws.SetCell(21, 3,'Hardware agreement (Hardware)',fmt4)
    self.ws.SetCell(21, 4,'=SUM(GS202:HE202)',fmt4b)
    self.ws.SetCell(21, 5,'=D21/SUM(D20:D27)*100',fmt4b)
    self.ws.SetCell(22, 3,'Post Sales support (SW-Customer Funded)',fmt4)
    self.ws.SetCell(22, 4,'=SUM(GS205:HE205)',fmt4b)
    self.ws.SetCell(22, 5,'=D22/SUM(D20:D27)*100',fmt4b)
    self.ws.SetCell(23, 3,'Post Sales support (HW-Customer Funded',fmt4)
    self.ws.SetCell(23, 4,'=SUM(GS206:HE206)',fmt4b)
    self.ws.SetCell(23, 5,'=D23/SUM(D20:D27)*100',fmt4b)
    self.ws.SetCell(24, 3,'NRE (Customer funded)',fmt4)
    self.ws.SetCell(24, 4,'=SUM(GS207:HE207)',fmt4b)
    self.ws.SetCell(24, 5,'=D24/SUM(D20:D27)*100',fmt4b)
    self.ws.SetCell(25, 3,'Training - Providing - Non Customer Specific',fmt4)
    self.ws.SetCell(25, 4,'=SUM(GS208:HE208)',fmt4b)
    self.ws.SetCell(25, 5,'=D25/SUM(D20:D27)*100',fmt4b)
    self.ws.SetCell(26, 3,'Training - Providing - Customer Specific',fmt4)
    self.ws.SetCell(26, 4,'=SUM(GS209:HE209)',fmt4b)
    self.ws.SetCell(26, 5,'=D26/SUM(D20:D27)*100',fmt4b)
    self.ws.SetCell(27, 3,'Post Sales Support (Warranty Period)',fmt4)
    self.ws.SetCell(27, 4,'=SUM(GS214:HE214)',fmt4b)
    self.ws.SetCell(27, 5,'=D27/SUM(D20:D27)*100',fmt4b)
    self.ws.SetCell(28, 3,'Pre Sales Support',fmt5)
    self.ws.SetCell(28, 4,'=SUM(GS203:HE203)',fmt5b)
    self.ws.SetCell(28, 5,'=D28/SUM(D28:D33)*100',fmt5b)
    self.ws.SetCell(29, 3,'Post Sales Support (Non contract)',fmt5)
    self.ws.SetCell(29, 4,'=SUM(GS204:HE204)',fmt5b)
    self.ws.SetCell(29, 5,'=D29/SUM(D28:D33)*100',fmt5b)
    self.ws.SetCell(30, 3,'Training - Receiving - Non Customer Specific',fmt5)
    self.ws.SetCell(30, 4,'=SUM(GS210:HE210)',fmt5b)
    self.ws.SetCell(30, 5,'=D30/SUM(D28:D33)*100',fmt5b)
    self.ws.SetCell(31, 3,'Training - Receiving - Customer Specific',fmt5)
    self.ws.SetCell(31, 4,'=SUM(GS211:HE211)',fmt5b)
    self.ws.SetCell(31, 5,'=D31/SUM(D28:D33)*100',fmt5b)
    self.ws.SetCell(32, 3,'Internal Business Meeting',fmt5)
    self.ws.SetCell(32, 4,'=SUM(GS212:HE212)',fmt5b)
    self.ws.SetCell(32, 5,'=D32/SUM(D28:D33)*100',fmt5b)
    self.ws.SetCell(33, 3,'Professional Services',fmt5)
    self.ws.SetCell(33, 4,'=SUM(GS213:HE213)',fmt5b)
    self.ws.SetCell(33, 5,'=D33/SUM(D28:D33)*100',fmt5b)
    self.ws.SetCell(34, 3,'Downtime - Excluding Leave & Sickness',fmt6)
    self.ws.SetCell(34, 4,'=SUM(GS232:HE232)',fmt6b)
    self.ws.SetCell(34, 5,'=D34/SUM(D34:D34)*100',fmt6b)
    self.ws.SetCell(35, 3,'Leave and Sickness',fmt7)
    self.ws.SetCell(35, 4,'=SUM(GS237:HE237)',fmt7b)
    self.ws.SetCell(35, 5,'=D35/SUM(D35:D35)*100',fmt7b)

    self.ws.SetCell(37, 3,'Customer split',fmt2)
    self.ws.SetCell(37, 4,'Hours',fmt2a)
    self.ws.SetCell(37, 5,'As a % of Total',fmt2a)
    self.ws.SetCell(38, 3,'Ericsson',fmt3)
    self.ws.SetCell(38, 4,'=SUM(GS247:HE247)',fmt3b)
    self.ws.SetCell(38, 5,'=D38/SUM(D38:D44)*100',fmt3b)
    self.ws.SetCell(39, 3,'Nokia',fmt3)
    self.ws.SetCell(39, 4,'=SUM(GS248:HE248)',fmt3b)
    self.ws.SetCell(39, 5,'=D39/SUM(D38:D44)*100',fmt3b)
    self.ws.SetCell(40, 3,'Alcatel-Lucent',fmt3)
    self.ws.SetCell(40, 4,'=SUM(GS249:HE249)',fmt3b)
    self.ws.SetCell(40, 5,'=D40/SUM(D38:D44)*100',fmt3b)
    self.ws.SetCell(41, 3,'Sum of all other customers',fmt3)
    self.ws.SetCell(41, 4,'=SUM(GS250:HE250)',fmt3b)
    self.ws.SetCell(41, 5,'=D41/SUM(D38:D44)*100',fmt3b)
    self.ws.SetCell(42, 3,'Cobham',fmt3)
    self.ws.SetCell(42, 4,'=SUM(GS251:HE251)',fmt3b)
    self.ws.SetCell(42, 5,'=D42/SUM(D38:D44)*100',fmt3b)
    self.ws.SetCell(43, 3,'Technical Training - All Types',fmt3)
    self.ws.SetCell(43, 4,'=SUM(GS252:HE251)',fmt3b)
    self.ws.SetCell(43, 5,'=D43/SUM(D38:D44)*100',fmt3b)
    self.ws.SetCell(44, 3,'Customer \'Other\'',fmt3)
    self.ws.SetCell(44, 4,'=SUM(GS253:HE251)',fmt3b)
    self.ws.SetCell(44, 5,'=D44/SUM(D38:D44)*100',fmt3b)

    self.ws.SetCell(46, 3,'Labour and Travel',fmt2)
    self.ws.SetCell(46, 4,'Hours',fmt2a)
    self.ws.SetCell(46, 5,'As a % of Total',fmt2a)
    self.ws.SetCell(47, 3,'Number of Labour Hours',fmt3)
    self.ws.SetCell(47, 4,'=SUM(GS217:HE217)',fmt3b)
    self.ws.SetCell(47, 5,'=D47/SUM(D47:D49)*100',fmt3b)
    self.ws.SetCell(48, 3,'Number of Labour Hours',fmt3)
    self.ws.SetCell(48, 4,'=SUM(GS218:HE218)',fmt3b)
    self.ws.SetCell(48, 5,'=D48/SUM(D47:D49)*100',fmt3b)
    self.ws.SetCell(49, 3,'Number of Labour Hours',fmt3)
    self.ws.SetCell(49, 4,'=SUM(GS219:HE219)',fmt3b)
    self.ws.SetCell(49, 5,'=D49/SUM(D47:D49)*100',fmt3b)

    self.ws.DrawBorder( 6, 3, 10, 5,'medium')
    self.ws.DrawBorder( 6, 3,  6, 5,'medium')
    self.ws.DrawBorder(12, 3, 16, 5,'medium')
    self.ws.DrawBorder(12, 3, 12, 5,'medium')
    self.ws.DrawBorder(19, 3, 35, 5,'medium')
    self.ws.DrawBorder(19, 3, 19, 5,'medium')
    self.ws.DrawBorder(37, 3, 44, 5,'medium')
    self.ws.DrawBorder(37, 3, 37, 5,'medium')
    self.ws.DrawBorder(46, 3, 49, 5,'medium')
    self.ws.DrawBorder(46, 3, 46, 5,'medium')


    ltsChart = PieChart()
    data   = Reference(self.ws.ws,min_col=5,min_row=47,max_row=49)
    labels = Reference(self.ws.ws,min_col=3,min_row=47,max_row=49)
    ltsChart.add_data(data)
    ltsChart.set_categories(labels)
    ltsChart.title = 'Labour Vs Travel'
    self.ws.ws.add_chart(ltsChart,'H46')


