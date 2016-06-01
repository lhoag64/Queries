import datetime
import re
import os.path
import logging
from   openpyxl import load_workbook
 
#-----------------------------------------------------------------------
# Class Timesheet Entry
#-----------------------------------------------------------------------
class TsEntry:
  def __init__(self,date,partA,partB,partC,partD,hours,workType,note,file,tsname,tsdate):
    self.date      = date
    self.code      = partA
    self.location  = partB
    self.activity  = partC
    self.product   = partD
    self.hours     = hours
    self.workType  = workType
    self.note      = note
    self.tsfile    = os.path.basename(file)
    self.tsname    = tsname
    self.tsdate    = tsdate

#-----------------------------------------------------------------------
# Class Timesheet
#-----------------------------------------------------------------------
class Timesheet:
  def __init__(self,tsInfo,wsDate):
    self.entries = []
    self.tsInfo  = tsInfo
    self.wsDate  = wsDate

  #---------------------------------------------------------------------
  def calcDate(self,day):
    day = day.lower().strip()
    day = day[0:3]
    day = { 'mon':0,'tue':1,'wed':2,'thu':3,'fri':4,'sat':5,'sun':6 }.get(day,7)
    if (day == 7):
      raise Exception
    date = self.wsDate + datetime.timedelta(days = day)
    return date

  #---------------------------------------------------------------------
#  def Clean(self):
#    for i,entry in enumerate(self.entries):
#      entry.Log()

  #---------------------------------------------------------------------
  def ReadFile(self):
    path = self.tsInfo.filename
    wb = load_workbook(path)
    sheet_names = wb.get_sheet_names();
    if (sheet_names[0] != 'Timesheet'):
      logging.error(path + ' is not a valid Timesheet spreadsheet')

    try:
      ws = wb.get_sheet_by_name('Timesheet')
    except KeyError:
      logging.error('Can\' find worksheet Timesheet in ' + path)
      return

    row = None
    col = None
    for rowIdx in range(1,9):
      for colIdx in range(1,9):
        val = ws.cell(row=rowIdx,column=colIdx).value
        if (type(val) is str):
          val = val.upper()
          #logging.debug(str(rowIdx).rjust(1) + '|' + str(colIdx).rjust(1) + '|' + val)
          match = re.search('CUSTOMER...PART A',val)
          if (match != None):
            row = rowIdx
            col = colIdx
            break

    if (row != None and col != None):
      startRow = row + 1
      startCol = col - 1
      #logging.debug('row: ' + str(startRow).rjust(1) + '|' + 'col: ' + str(startCol).rjust(1))
    else:
      logging.error('Couldn\'t sync to Timesheet spreadsheet')
      return

    row = None
    col = None
    for rowIdx in range(1,9):
      for colIdx in range(1,9):
        val = ws.cell(row=rowIdx,column=colIdx).value
        if (type(val) is str):
          val = val.upper()
          #logging.debug(str(rowIdx).rjust(1) + '|' + str(colIdx).rjust(1) + '|' + val)
          match = re.search('NAME:',val)
          if (match != None):
            row = rowIdx
            col = colIdx
            break

    if (row != None and col != None):
      nameFromFile = str(ws.cell(row=row,column=col+2).value).strip()
      if (nameFromFile[0:4] == 'Week'):
        nameFromFile = str(ws.cell(row=row,column=col+1).value).strip()
    else:
      nameFromFile = ''

    row = None
    col = None
    for rowIdx in range(1,9):
      for colIdx in range(1,9):
        val = ws.cell(row=rowIdx,column=colIdx).value
        if (type(val) is str):
          val = val.upper()
          #logging.debug(str(rowIdx).rjust(1) + '|' + str(colIdx).rjust(1) + '|' + val)
          match = re.search('WEEK COMMENCING',val)
          if (match != None):
            row = rowIdx
            col = colIdx
            break

    if (row != None and col != None):
      dateFromFile = str(ws.cell(row=row,column=col+1).value).strip()
      dateFromFile = dateFromFile[0:10]
    else:
      dateFromFile = ''

    tsname = nameFromFile
    tsdate = dateFromFile

    #nameFromFile = nameFromFile.ljust(18)
    #dateFromFile = dateFromFile.ljust(15)

    logging.debug(tsname.ljust(30) + '|' + tsdate.ljust(10) + '|Reading ' + path)

    wsRow     = startRow
    wsCol     = startCol
    sundayFlg = False
    sundayCnt = 0;
    blankFlg  = False
    curDate   = None

    while True:
      day = ws.cell(row=wsRow,column=wsCol+0).value

      cell = ws.cell(row=wsRow,column=wsCol+0)
      fill = cell.fill
      if (fill.patternType == 'solid'):
        if (sundayFlg):
          #logging.debug('Giving up due to fill pattern')
          break

      if (day == None):
        day = ''
        if (curDate == None):
          logging.error('SHOULDNT BE HERE')
      else:
        curDate = self.calcDate(day)
 
      partA  = ws.cell(row=wsRow,column=wsCol+1).value
      if (partA == None):
        partA = ''
        blankFlg = True

      partB = ws.cell(row=wsRow,column=wsCol+2).value
      if (partB == None):
        partB = ''

      partC = ws.cell(row=wsRow,column=wsCol+3).value
      if (partC == None):
        partC = ''

      partD = ws.cell(row=wsRow,column=wsCol+4).value
      if (partD == None):
        partD = ''

      hours = ws.cell(row=wsRow,column=wsCol+9).value
      if (hours == None):
        hours = ''
      else:
        try:
          hours = float(hours)
        except:
          hours = ''

      ltd = ws.cell(row=wsRow,column=wsCol+10).value
      if (ltd == None):
        ltd = ''

      notes = ws.cell(row=wsRow,column=wsCol+11).value
      if (notes == None):
        notes = ''
 
      # TODO: if blankFlg == False, then make sure everything else is blank

      if (not blankFlg):
        self.entries.append(TsEntry(curDate,partA,partB,partC,partD,hours,ltd,notes,path,tsname,tsdate))

      if (day.startswith('Sunday')):
        sundayFlg = True
      if (sundayFlg):
        if (blankFlg):
          sundayCnt += 1
        else:
          sundayCnt = 0;
      if (sundayCnt > 3):
        break


      wsRow += 1
      blankFlg = False

  #---------------------------------------------------------------------
#  def Log(self):
#    for i,entry in enumerate(self.entries):
#      fullname = self.tsInfo.fae.fullname.ljust(25)
#      date  = str(self.entries[i].date)
#      partA = str(self.entries[i].code).ljust(50)
#      partB = str(self.entries[i].location).ljust(60)
#      partC = str(self.entries[i].activity).ljust(45)
#      partD = str(self.entries[i].product).ljust(35)
#      hours = str(self.entries[i].hours).ljust(5)
#      ltd   = str(self.entries[i].worktype).ljust(10)
#      logging.debug(fullname + '|' + str(date) + '|' + partA + '|' + partB + '|' + partC + '|' + partD + '|' + hours + '|' + ltd)

