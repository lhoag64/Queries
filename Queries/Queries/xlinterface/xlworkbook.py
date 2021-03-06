import logging
import openpyxl
from   openpyxl                import load_workbook
from   openpyxl.workbook       import Workbook
from   xlinterface.xlworksheet import XlWorkSheet

#----------------------------------------------------------------------
class XlWorkBook:
  def __init__(self):
    self.wsByName = {}
    self.wsByPyWs = {}
    self.wsByXlWs = {}

    self.wb = Workbook()
    pyws    = self.wb.active
    name    = pyws.title
    xlws = XlWorkSheet(self.wb,pyws,name)
    self.wsByName[name] = (pyws,xlws)
    self.wsByPyWs[pyws] = (xlws,name)
    self.wsByXlWs[xlws] = (pyws,name)

    self.activeSheet = name

  #--------------------------------------------------------------------
  def Read(self,filename):
    self.wsByName = {}
    self.wsByPyWs = {}
    self.wsByXlWs = {}

    self.wb = load_workbook(filename)

    for item in self.wb.worksheets:
      pyws = item
      name = item.title
      xlws = XlWorkSheet(self.wb,pyws,name)
      self.wsByName[name] = (pyws,xlws)
      self.wsByPyWs[pyws] = (xlws,name)
      self.wsByXlWs[xlws] = (pyws,name)

    self.activeSheet = self.wb.active.title

  #--------------------------------------------------------------------
  def CreateXlWorkSheet(self,name):
    if (name not in self.wsByName):
      self.wb.create_sheet(name)
      pyws = self.wb.get_sheet_by_name(name)
      xlws = XlWorkSheet(self.wb,pyws,name)

      self.wsByName[name] = (pyws,xlws)
      self.wsByPyWs[pyws] = (xlws,name)
      self.wsByXlWs[xlws] = (pyws,name)
    
      return self.wsByName[name][1]

    else:
      logging.error('Attempting to create duplicate name in workbook: ' + name)
      return None

  #--------------------------------------------------------------------
  def CreateXlChrtSheet(self,name):
    if (name not in self.wsByName):
      pycs = self.wb.create_chartsheet(name)
      #pycs = self.wb.get_sheet_by_name(name)
      xlcs = XlChrtSheet(self.wb,pycs,name)

      self.wsByName[name] = (pycs,xlcs)
      #self.wsByPyWs[pycs] = (xlcs,name)
      self.wsByXlWs[xlcs] = (pycs,name)
    
      return self.wsByName[name][1]

    else:
      logging.debug.error('Attempting to create duplicate name in workbook: ' + name)
      return None

  #--------------------------------------------------------------------
#  def CreateSheet(self,name):
#    if (name not in self.wsByName):
#      self.wb.create_sheet(name)
#      pyws = self.wb.get_sheet_by_name(name)
#      xlws = XlWorkSheet(self.wb,pyws,name)
#
#      self.wsByName[name] = (pyws,xlws)
#      self.wsByPyWs[pyws] = (xlws,name)
#      self.wsByXlWs[xlws] = (pyws,name)
#    
#      return self.wsByName[name][1]
#
#    else:
#      debug.error('Attempting to create duplicate name in workbook: ' + name)
#      return None

  #--------------------------------------------------------------------
  def GetActiveSheet(self):
    pyws = self.wb.active
    name = pyws.title
    if (name in self.wsByName):
      if (name == self.activeSheet):
        return self.wsByName[name][1]
      else:
        logging.debug('XlWorkBook data corrupt')
    else:
      logging.debug('XlWorkBook data corrupt')
  
  #--------------------------------------------------------------------
  def SetName(self,ws,name):
    if (name not in self.wsByName):
      xlws = ws
      pyws,orig = self.wsByXlWs[xlws]
      pyws.title = name
      xlws.name  = name

      self.wsByPyWs[pyws] = (xlws,name)
      self.wsByXlWs[xlws] = (pyws,name)

      if (self.activeSheet == orig):
        self.activeSheet = name

      del self.wsByName[orig]
      self.wsByName[name] = (pyws,xlws)

    else:
      debug.error('Attempting to create duplicate name in workbook: ' + name)

    return self.wsByName[name][1]

  #--------------------------------------------------------------------
  def GetSheetByName(self,name):
    return self.wsByName[name][1]

  #--------------------------------------------------------------------
  def RemoveSheetByName(self,name):
    self.wb.remove_sheet(self.wb.get_sheet_by_name(name))
  #--------------------------------------------------------------------
  def Save(self,name):
    self.wb.save(name)

