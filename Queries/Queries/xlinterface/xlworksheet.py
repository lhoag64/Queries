import logging
import openpyxl
from   openpyxl.styles                     import Font,Border,Alignment,Color,Style,PatternFill
from   openpyxl.styles.borders             import Side
from   openpyxl.styles.fills               import FILL_SOLID
from   xlinterface.xlcolortable            import XlColorTable as ColorTable
from   openpyxl.workbook.names.named_range import NamedRange

alignType = {'C':'center','L':'left','R':'right'}
colorType = {'Black':0,'Red':2,'Green':3,'Orange':19}

class XlWorkSheet:
  def __init__(self,wb,ws,name):
    self.wb   = wb
    self.ws   = ws
    self.name = name

  #-------------------------------------------------------------------
  def GetColumnLetter(self,index):
    return openpyxl.utils._get_column_letter(index)

  #-------------------------------------------------------------------
  def GetColumnIndex(self,letter):
    return openpyxl.utils.column_index_from_string(letter)

  #-------------------------------------------------------------------
  def SetColWid(self,col,wid):
    self.ws.column_dimensions[self.GetColumnLetter(col)].width = wid
  
  #-------------------------------------------------------------------
  def SetRowHgt(self,col,hgt):
    self.ws.row_dimensions[col].height = hgt

  #-------------------------------------------------------------------
  def DrawRegion(self,tRow,lCol,bRow,rCol,bType=None,color=None):
    ws = self.ws
    if (color):
      cFmt = {'fill':color}
      for i in range(lCol,rCol+1):
        for j in range(tRow,bRow+1):
          self.SetFormat(j,i,cFmt)
    if (bType):
      for i in range(lCol,rCol+1):
        tFmt = {'border':{'T':bType}}
        self.SetFormat(tRow,i,tFmt)
        bFmt = {'border':{'B':bType}}
        self.SetFormat(bRow,i,bFmt)
      for j in range(tRow,bRow+1):
        lFmt = {'border':{'L':bType}}
        self.SetFormat(j,lCol,lFmt)
        rFmt = {'border':{'R':bType}}
        self.SetFormat(j,rCol,rFmt)

  #-------------------------------------------------------------------
  def DrawBorder(self,tRow,lCol,bRow,rCol,type):
    ws = self.ws
    for i in range(lCol,rCol+1):
      tFmt = {'border':{'T':type}}
      self.SetFormat(tRow,i,tFmt)
      bFmt = {'border':{'B':type}}
      self.SetFormat(bRow,i,bFmt)
    for j in range(tRow,bRow+1):
      lFmt = {'border':{'L':type}}
      self.SetFormat(j,lCol,lFmt)
      rFmt = {'border':{'R':type}}
      self.SetFormat(j,rCol,rFmt)

  # Border Side Style
  #   hair
  #   thin
  #   medium
  #   thick

  #-------------------------------------------------------------------
  def SetFormat(self,row,col,fmt):
    ws     = self.ws
    font   = None
    color  = None
    align  = None
    fill   = None
    numFmt = None
    border = None

    c = ws.cell(row=row,column=col)

    #-------------------------------------------------------------------------
    for i in fmt:
      if (i == 'hAlign'): 
        if (not align): align = Alignment()
        align.horizontal = alignType[fmt[i]]
      elif (i == 'vAlign'): 
        if (not align): align = Alignment()
        align.vertical   = alignType[fmt[i]]
      elif (i == 'tAlign'): 
        if (not align): align = Alignment()
        align.text_rotation = fmt[i]
      elif (i == 'wrap'): 
        if (not align): align = Alignment()
        align.wrap_text = fmt[i]

      elif (i == 'font'):
        name = 'Calibri'
        bold = False
        size = 11
        dict = fmt[i]
        if ('emph' in dict):
          if (dict['emph'] == 'B'):
            bold = True
        if ('size' in dict):
          size = dict['size']
        if (not font):
            font = Font(name=name,size=size,bold=bold)

      elif (i == 'border'):
        dict = fmt[i]
        color = None
        style = None
        if ('Color' in dict):
          color = ColorTable[dict['Color']]
        else:
          color = ColorTable['Black']
        if ('Style' in dict):
          color = dict['Style']
        if (c.border.top.style == None):
          tSide = Side(color=color)
        else:
          tSide = c.border.top.copy()
        if (c.border.bottom.style == None):
          bSide = Side(color=color)
        else:
          bSide = c.border.bottom.copy()
        if (c.border.left.style == None):
          lSide = Side(color=color)
        else:
          lSide = c.border.left.copy()
        if (c.border.right.style == None):
          rSide = Side(color=color)
        else:
          rSide = c.border.right.copy()

        if ((len(dict) ==1) and ('A' in dict)):
          tSide.style = dict['A']
          bSide.style = dict['A']
          lSide.style = dict['A']
          rSide.style = dict['A']
        else:
          for j in dict:
            if (j == 'T'):
              tSide.style = dict[j]
            if (j == 'B'):
              bSide.style = dict[j]
            if (j == 'L'):
              lSide.style = dict[j]
            if (j == 'R'):
              rSide.style = dict[j]

        border = Border(left=lSide,right=rSide,top=tSide,bottom=bSide)

      elif (i == 'fill'): 
        color = ColorTable[fmt[i]]
        fill = PatternFill(start_color=color,end_color='FFFFFFFF',fill_type='solid')

      elif (i == 'orient'): 
        pass

      elif (i == 'numFmt'):
        numFmt = fmt[i]

    #-------------------------------------------------------------------------
    if (font):
      c.font = font.copy()

    if (align):
      c.alignment = align.copy()

    if (border):
      c.border = border.copy()

    if (fill):
      c.fill = fill.copy()

    if (numFmt):
      c.number_format = numFmt

  #-------------------------------------------------------------------
  def SetCell(self,row,col,val,fmt=None):
    ws = self.ws

    if (fmt):
      self.SetFormat(row,col,fmt)

    c = ws.cell(row=row,column=col)
    c.value = val

  #-------------------------------------------------------------------
  def SetCellFmt(self,cell,hAlign,vAlign,fmt,value):
    if (align == 'C'):
      align = Alignment(horizontal='center',vertical='center')
    elif (align == 'L'):
      align = Alignment(horizontal='left',vertical='center')
    elif (align == 'R'):
      align = Alignment(horizontal='right',vertical='center')
    else:
      align = Alignment(horizontal='right',vertical='center')

    side   = Side(style='thin')
    border = Border(left=side,right=side,top=side,bottom=side)

    if (fmt == 'F'):
      fmt = '0.00'
      cell.number_format = fmt
    cell.alignment     = align.copy()
    cell.border        = border.copy()
    cell.value = value

  #-------------------------------------------------------------------
  def GetValue(self,wsRow,wsCol):
    ws = self.ws
    c = ws.cell(row=wsRow,column=wsCol)
    return c.value

  #-------------------------------------------------------------------
  def GetCell(self,wsRow,wsCol):
    ws = self.ws
    c = ws.cell(row=wsRow,column=wsCol)
    return c.value

  #-------------------------------------------------------------------
  def Cell(self,wsRow,wsCol):
    ws = self.ws
    c = ws.cell(row=wsRow,column=wsCol)
    return c

  #-------------------------------------------------------------------
  def FlgCell(self,wsRow,wsCol):
    ws = self.ws
    c = ws.cell(row=wsRow,column=wsCol)

    red = openpyxl.styles.colors.RED
    side   = Side(style='medium',color=red)
    border = Border(left=side,right=side,top=side,bottom=side)
    c.border = border.copy()

  #-------------------------------------------------------------------
  def WsCell(self,wsRow,wsCol,val=None):
    ws = self.ws
    if val is not None:
      return ws.cell(row=wsRow,column=wsCol,value=value)
    else:
      return ws.cell(row=wsRow,column=wsCol)


  #-------------------------------------------------------------------
  def CopyCell(srcWs,srcRow,srcCol,dstWs,dstRow,dstCol):
    scell = srcWs.ws.cell(row=srcRow,column=srcCol)
    dcell = dstWs.ws.cell(row=dstRow,column=dstCol)

    dcell.value         = scell.value
    dcell.border        = scell.border.copy()
    dcell.alignment     = scell.alignment.copy()
    dcell.fill          = scell.fill.copy()
    dcell.font          = scell.font.copy()
    dcell.protection    = scell.protection.copy()
    dcell.number_format = scell.number_format
    dcell.data_type     = scell.data_type
    #dcell.is_date       = scell.is_date
    #dcell.has_style     = scell.has_style
    #dcell.style_id      = scell.style_id

  #-------------------------------------------------------------------
  def cleanRangeName(self,name):
    text = ''
    for i in name:
      if (i in [' ','-','\n','\r','\t']):
        ch = '_'
      else:
        ch = i
      text += ch
    return text

  #-------------------------------------------------------------------
  def AddNamedRange(self,name,sRow,sCol,eRow,eCol):

    logging.debug(name)

    name = self.cleanRangeName(name)

    pyWs = self.ws
    pyWb = self.wb

    sCol = self.GetColumnLetter(sCol)
    eCol = self.GetColumnLetter(eCol)

    range = '$' + sCol + '$' + str(sRow) + ':' + '$' + eCol + '$' + str(eRow)

    destination = [(pyWs,range)]
    scope       = None
    nr          = NamedRange(name,destination,scope)

    pyWb.add_named_range(nr)

    return nr
